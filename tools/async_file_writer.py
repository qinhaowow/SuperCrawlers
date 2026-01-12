#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (c) 2025 SuperCrawler Project
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1

import asyncio
import os
import json
import csv
import time
from typing import Dict, Optional, List, Any, Union
from datetime import datetime
import aiofiles

from config import base_config


class AsyncFileWriter:
    """Asynchronous file writer for SuperCrawler"""
    
    def __init__(self, 
                 platform: str, 
                 crawler_type: str,
                 output_file: Optional[str] = None):
        """Initialize async file writer"""
        self.platform = platform
        self.crawler_type = crawler_type
        self.output_file = output_file
        self.data_dir = base_config.DATA_DIR
        
        # Ensure data directory exists
        os.makedirs(self.data_dir, exist_ok=True)
        
        # Generate output file if not provided
        if not self.output_file:
            self.output_file = self._generate_output_file()
    
    def _generate_output_file(self) -> str:
        """Generate output file path based on platform and crawler type"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{self.platform}_{self.crawler_type}_{timestamp}"
        
        # Determine file extension based on save option
        if base_config.SAVE_DATA_OPTION == "json":
            file_name += ".json"
        elif base_config.SAVE_DATA_OPTION == "csv":
            file_name += ".csv"
        elif base_config.SAVE_DATA_OPTION == "excel":
            file_name += ".xlsx"
        else:
            file_name += ".json"  # Default to JSON
        
        return os.path.join(self.data_dir, file_name)
    
    async def write_data(self, data: Union[List[Dict], Dict]) -> str:
        """Write data to file asynchronously"""
        start_time = time.time()
        print(f"[FileWriter] Writing data to {self.output_file}")
        
        try:
            if base_config.SAVE_DATA_OPTION == "json":
                await self._write_json(data)
            elif base_config.SAVE_DATA_OPTION == "csv":
                await self._write_csv(data)
            elif base_config.SAVE_DATA_OPTION == "excel":
                await self._write_excel(data)
            else:
                # Default to JSON
                await self._write_json(data)
            
            elapsed = time.time() - start_time
            print(f"[FileWriter] Data written successfully in {elapsed:.2f}s")
            return self.output_file
            
        except Exception as e:
            print(f"[FileWriter] Error writing data: {e}")
            raise
    
    async def _write_json(self, data: Union[List[Dict], Dict]):
        """Write data to JSON file"""
        async with aiofiles.open(self.output_file, "w", encoding="utf-8") as f:
            await f.write(json.dumps(data, ensure_ascii=False, indent=2))
    
    async def _write_csv(self, data: Union[List[Dict], Dict]):
        """Write data to CSV file"""
        # Convert single dict to list
        if isinstance(data, dict):
            data = [data]
        
        if not data:
            return
        
        # Get fieldnames from first item
        fieldnames = list(data[0].keys())
        
        async with aiofiles.open(self.output_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            await writer.writeheader()
            
            for item in data:
                # Convert non-string values to strings
                row = {k: str(v) if v is not None else "" for k, v in item.items()}
                await writer.writerow(row)
    
    async def _write_excel(self, data: Union[List[Dict], Dict]):
        """Write data to Excel file"""
        # Excel writing requires openpyxl
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            print("[FileWriter] openpyxl not installed, falling back to JSON")
            await self._write_json(data)
            return
        
        # Create workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"{self.platform}_{self.crawler_type}"
        
        # Convert single dict to list
        if isinstance(data, dict):
            data = [data]
        
        if data:
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Write data
            for row_idx, item in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save workbook
        workbook.save(self.output_file)
    
    async def generate_wordcloud_from_comments(self, comments_file: Optional[str] = None) -> str:
        """Generate wordcloud from comments"""
        try:
            from wordcloud import WordCloud
            import matplotlib.pyplot as plt
            import jieba
        except ImportError:
            print("[FileWriter] wordcloud or jieba not installed, skipping wordcloud generation")
            return ""
        
        # Use current output file if no comments file provided
        target_file = comments_file or self.output_file
        
        if not os.path.exists(target_file):
            print(f"[FileWriter] Comments file not found: {target_file}")
            return ""
        
        # Read comments from file
        comments = await self._read_comments(target_file)
        if not comments:
            print("[FileWriter] No comments found for wordcloud generation")
            return ""
        
        # Generate wordcloud
        wordcloud_file = await self._generate_wordcloud(comments)
        print(f"[FileWriter] Wordcloud generated: {wordcloud_file}")
        return wordcloud_file
    
    async def _read_comments(self, file_path: str) -> List[str]:
        """Read comments from file"""
        comments = []
        
        if file_path.endswith(".json"):
            async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
                data = json.loads(await f.read())
                if isinstance(data, list):
                    for item in data:
                        if "comment" in item:
                            comments.append(str(item["comment"]))
                        elif "content" in item:
                            comments.append(str(item["content"]))
                elif isinstance(data, dict):
                    if "comments" in data and isinstance(data["comments"], list):
                        for comment in data["comments"]:
                            if "content" in comment:
                                comments.append(str(comment["content"]))
                            elif "text" in comment:
                                comments.append(str(comment["text"]))
        
        elif file_path.endswith(".csv"):
            async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
                content = await f.read()
                reader = csv.DictReader(content.splitlines())
                for row in reader:
                    if "comment" in row:
                        comments.append(row["comment"])
                    elif "content" in row:
                        comments.append(row["content"])
        
        return comments
    
    async def _generate_wordcloud(self, comments: List[str]) -> str:
        """Generate wordcloud from comments"""
        # Combine all comments
        text = " ".join(comments)
        
        # Tokenize Chinese text
        if self.platform in ["xhs", "dy", "ks", "bili", "wb", "tieba", "zhihu"]:
            text = " ".join(jieba.cut(text))
        
        # Generate wordcloud
        wordcloud = WordCloud(
            width=1200,
            height=800,
            background_color="white",
            font_path=self._get_font_path(),
            max_words=200,
            max_font_size=100,
            random_state=42
        ).generate(text)
        
        # Save wordcloud
        wordcloud_file = self.output_file.replace(".json", "_wordcloud.png")
        wordcloud_file = wordcloud_file.replace(".csv", "_wordcloud.png")
        wordcloud_file = wordcloud_file.replace(".xlsx", "_wordcloud.png")
        
        # Save to file
        wordcloud.to_file(wordcloud_file)
        
        return wordcloud_file
    
    def _get_font_path(self) -> Optional[str]:
        """Get font path for Chinese text"""
        # Try to find a Chinese font
        font_paths = [
            "C:/Windows/Fonts/simhei.ttf",  # Windows
            "C:/Windows/Fonts/msyh.ttf",    # Windows
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",  # Linux
            "/Library/Fonts/Songti.ttc",    # macOS
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                return font_path
        
        return None  # Use default font
    
    async def append_data(self, data: Union[List[Dict], Dict]) -> str:
        """Append data to existing file"""
        if not os.path.exists(self.output_file):
            return await self.write_data(data)
        
        print(f"[FileWriter] Appending data to {self.output_file}")
        
        try:
            if self.output_file.endswith(".json"):
                await self._append_json(data)
            elif self.output_file.endswith(".csv"):
                await self._append_csv(data)
            else:
                # For other formats, rewrite the entire file
                existing_data = await self._read_data()
                if isinstance(existing_data, list):
                    if isinstance(data, list):
                        existing_data.extend(data)
                    else:
                        existing_data.append(data)
                    await self.write_data(existing_data)
                else:
                    await self.write_data(data)
            
            return self.output_file
            
        except Exception as e:
            print(f"[FileWriter] Error appending data: {e}")
            raise
    
    async def _read_data(self) -> Union[List[Dict], Dict]:
        """Read data from file"""
        if not os.path.exists(self.output_file):
            return []
        
        if self.output_file.endswith(".json"):
            async with aiofiles.open(self.output_file, "r", encoding="utf-8") as f:
                return json.loads(await f.read())
        
        return []
    
    async def _append_json(self, data: Union[List[Dict], Dict]):
        """Append data to JSON file"""
        existing_data = await self._read_data()
        
        if isinstance(existing_data, list):
            if isinstance(data, list):
                existing_data.extend(data)
            else:
                existing_data.append(data)
        else:
            existing_data = [existing_data, data]
        
        await self._write_json(existing_data)
    
    async def _append_csv(self, data: Union[List[Dict], Dict]):
        """Append data to CSV file"""
        # Convert single dict to list
        if isinstance(data, dict):
            data = [data]
        
        if not data:
            return
        
        # Check if file exists and has headers
        has_headers = os.path.exists(self.output_file) and os.path.getsize(self.output_file) > 0
        
        async with aiofiles.open(self.output_file, "a", encoding="utf-8", newline="") as f:
            if data:
                fieldnames = list(data[0].keys())
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                
                # Write headers only if file is empty
                if not has_headers:
                    await writer.writeheader()
                
                # Write data
                for item in data:
                    row = {k: str(v) if v is not None else "" for k, v in item.items()}
                    await writer.writerow(row)


async def write_data_async(platform: str, 
                           crawler_type: str, 
                           data: Union[List[Dict], Dict],
                           output_file: Optional[str] = None) -> str:
    """Helper function to write data asynchronously"""
    writer = AsyncFileWriter(platform=platform, crawler_type=crawler_type, output_file=output_file)
    return await writer.write_data(data)


async def generate_wordcloud_async(platform: str, 
                                  crawler_type: str,
                                  comments_file: Optional[str] = None) -> str:
    """Helper function to generate wordcloud asynchronously"""
    writer = AsyncFileWriter(platform=platform, crawler_type=crawler_type)
    return await writer.generate_wordcloud_from_comments(comments_file)